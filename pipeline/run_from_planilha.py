# -*- coding: utf-8 -*-
"""Le Pole/FILA DE TRABALHO/links_pendentes.xlsx (coluna 'link', coluna 'Processado?' com
'X' nas ja feitas), processa so as linhas sem X, escreve de volta cod_id /
endereco / Processado=X / precisa_confirmar / motivo em cada linha processada,
e salva a planilha.

Convencao de nomes de arquivo (definida pelo usuario 23/ago): o nome do
arquivo em si NUNCA carrega comentario/motivo -- so `COD ID_<cod_id>.jpg` ou
`COD ID_<cod_id>!!.jpg` quando precisa confirmar. Todo comentario vai pra um
relatorio a parte:
  - `RELATORIOS/confirmar_manual.xlsx` -- flagueados que NAO sao duplicata
    (distancia/rumo nao bateram com confianca), com link + motivo.
  - `RELATORIOS/duplicados_para_validar.xlsx` -- toda vez que dois links
    diferentes resolvem pro MESMO cod_id, nenhum dos dois fica na fila normal:
    ambos (o antigo, movido, e o novo) vao pra pasta `POSTES UTILIZADOS/DUPLICADOS/`,
    nomeados por linha da planilha (`COD ID_<cod_id> (linha N).jpg`), e essa
    planilha lista os links de cada um lado a lado pra decidir qual e o certo.
"""
import re
import os
import sys
import csv
import difflib
import glob
import math
import shutil
import time
import json

import numpy as np
import cv2
import openpyxl
from openpyxl.styles import Font
from PIL import Image
from playwright.sync_api import sync_playwright

from pano import get_pano_meta, normalize_street_name, yaw_from_bearing
from link_cleanup import remove_watermark_v3

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLANILHA = os.path.join(ROOT, "FILA DE TRABALHO", "links_pendentes.xlsx")
POSTES_CSV = os.path.join(ROOT, "BASE DE DADOS", "lista_postes.csv")
ESQ_CSV = os.path.join(ROOT, "BASE DE DADOS", "esquina_index.csv")
DUPLICADOS_DIR = os.path.join(ROOT, "POSTES UTILIZADOS", "DUPLICADOS")
DUPLICADOS_XLSX = os.path.join(ROOT, "RELATORIOS", "duplicados_para_validar.xlsx")
CONFIRMAR_XLSX = os.path.join(ROOT, "RELATORIOS", "confirmar_manual.xlsx")
PROCESSADOS_XLSX = os.path.join(ROOT, "RELATORIOS", "links_postes_processados.xlsx")

VIEWPORT = {"width": 2547, "height": 1532}
TOP_MARGIN, BOTTOM_MARGIN = 175, 145
LEFT_UI_EXCLUDE = 350
RIGHT_UI_EXCLUDE = 2547 - 120
OUT_W, OUT_H = 1000, 1333

RE_MAIN = re.compile(r"@(-?\d+\.\d+),(-?\d+\.\d+),\d+a,([\d.]+)y,([\d.]+)h,([\d.]+)t")
# o "3a" logo apos lat/lon e' so um marcador fixo do Google (modo Street
# View), nao carrega nenhum dado geografico -- aceitar \d+a em vez de so
# "3a" (achado 26/ago: 210 linhas vieram com esse numero virando 959a/960a
# etc., cara de arrasto de preenchimento automatico do Excel incrementando
# um numero dentro do texto sem querer; o resto do link -- lat/lon/y/h/t/
# panoid -- continua intacto e confiavel, so esse marcador que mudou)
RE_PANOID = re.compile(r"!1s([A-Za-z0-9_-]{20,30})!2e\d+")
RE_COORDS_ONLY = re.compile(r"@(-?\d+\.\d+),(-?\d+\.\d+)")
RE_VIEWPOINT = re.compile(r"viewpoint=(-?\d+\.\d+),(-?\d+\.\d+)")


def extrai_coords_fallback(candidatos):
    """Procura coordenada em qualquer um dos candidatos (valor direto ou
    hyperlink), nao so no primeiro -- e aceita tanto '@lat,lon' quanto
    'viewpoint=lat,lon' (formato do link "generico" sem angulo, tipo o que
    lista_postes.csv guarda em street_view_url). Achado 26/ago."""
    for cand in candidatos:
        if not cand:
            continue
        m = RE_COORDS_ONLY.search(cand) or RE_VIEWPOINT.search(cand)
        if m:
            return float(m.group(1)), float(m.group(2))
    return None


def link_candidates(cell):
    """O Excel as vezes autolinca um texto colado (tipo o titulo da aba do
    navegador, "436 R. Sao Luis - Google Maps") como hyperlink, guardando o
    link de verdade em cell.hyperlink.target -- cell.value fica so com o
    rotulo visivel, que o parse_link nunca reconhece. Achado 25/ago: 30
    linhas na planilha tinham exatamente esse caso e ficavam marcadas "URL
    nao reconhecida" pra sempre.

    Em vez de escolher um dos dois cegamente, devolve as duas fontes
    possiveis (valor direto primeiro, "caminho tradicional") pra quem chama
    tentar cada uma e ficar com a que der certo -- pedido do usuario 25/ago,
    pra nunca perder um link valido so por causa de qual caminho ele veio."""
    candidates = []
    if cell.value:
        candidates.append(cell.value)
    if cell.hyperlink and cell.hyperlink.target and cell.hyperlink.target not in candidates:
        candidates.append(cell.hyperlink.target)
    return candidates


def cell_url(cell):
    """Primeiro candidato disponivel (valor direto, senao hyperlink) -- usado
    so onde a gente so precisa saber SE existe algum link na celula, nao
    qual dos dois efetivamente parseia certo (isso e' link_candidates)."""
    c = link_candidates(cell)
    return c[0] if c else None


def parse_link(url):
    m = RE_MAIN.search(url)
    p = RE_PANOID.search(url)
    if not m or not p:
        return None
    lat, lon, y, h, t = m.groups()
    return dict(url=url, lat=float(lat), lon=float(lon), y=float(y), h=float(h),
                t=float(t), pitch=90 - float(t), panoid=p.group(1))


def hav(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def bearing_deg(lat1, lon1, lat2, lon2):
    la1, la2 = math.radians(lat1), math.radians(lat2)
    dl = math.radians(lon2 - lon1)
    x = math.sin(dl) * math.cos(la2)
    y = math.cos(la1) * math.sin(la2) - math.sin(la1) * math.cos(la2) * math.cos(dl)
    return math.degrees(math.atan2(x, y)) % 360


def ang_diff(a, b):
    d = abs(a - b) % 360
    return min(d, 360 - d)


NO_STREET = {"", "ARAXA", "ARAXA STATE OF MINAS GERAIS", "MINAS GERAIS"}


def street_word_similarity(a, b):
    """Similaridade por PALAVRA, nao por caractere -- mais robusta pra nome
    de rua (achado 23/ago: comparacao por caractere pura, tipo
    difflib.ratio() direto na string inteira, deu quase empatado - 0.35 -
    entre 'MARIA MULLER' e 'PEDRO DIAS DE CARVALHO', ruas completamente
    diferentes, so por coincidencia de letras soltas). Cada palavra da rua
    mais curta procura a mais parecida na outra rua (fuzzy OU uma comeca
    com a outra, pra pegar abreviacao tipo 'VER' de 'VEREADOR'); resultado e
    a fracao de palavras que acharam par razoavel."""
    wa, wb = a.split(), b.split()
    if not wa or not wb:
        return 0.0
    short, long_ = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    matched = 0
    for w in short:
        best = max((difflib.SequenceMatcher(None, w, w2).ratio() for w2 in long_), default=0.0)
        abbrev = any(len(w) >= 2 and (w2.startswith(w) or w.startswith(w2)) for w2 in long_)
        if best > 0.75 or abbrev:
            matched += 1
    return matched / len(short)


def load_poles():
    rows = []
    with open(POSTES_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            if row["poste_utilizado"] != "True":
                continue
            try:
                la, lo = float(row["lat"]), float(row["lon"])
            except Exception:
                continue
            rows.append((row["cod_id"], la, lo, row["endereco_lote_mais_proximo"]))
    return rows


def load_poste_numero_map():
    """poste_numero (o numero de campo/projeto, coluna 'Poste' que o usuario
    preenche a mao pra desambiguar esquina/postes proximos -- pedido
    23/ago) -> cod_id. E 1:1, unico pros 3.835 postes utilizados (conferido
    23/ago) -- quando preenchido, e a fonte da verdade e substitui o match
    por distancia/rumo/rua inteiro, nao so mais um sinal."""
    m = {}
    with open(POSTES_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            if row["poste_numero"]:
                m[row["poste_numero"].strip()] = row["cod_id"]
    return m


def load_base_links():
    """cod_id -> street_view_url oficial (gerado direto da coordenada
    cadastral, sem heading/pitch escolhido) -- usado como link de referencia
    pra validar visualmente se o link mirado bateu no poste certo."""
    links = {}
    with open(POSTES_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            links[row["cod_id"]] = row["street_view_url"]
    return links


def load_esquina():
    ids = set()
    with open(ESQ_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            ids.add(row["cod_id"])
    return ids


# cada categoria tem uma fila de revisao E uma pasta ja aprovada ("- OK"),
# pra nao misturar pendente com confirmado -- ver pipeline/routing.py
_CATEGORY_SUBDIRS = (
    "ESQUINA - REVISAO", "ESQUINA - OK",
    "POSTE UTILIZADO - REVISAO", "POSTE UTILIZADO - OK",
    "POSTE CASA - REVISAO", "POSTE CASA - OK",
    "BAIXA CONFIANCA - REVISAO", "BAIXA CONFIANCA - OK",
)


def already_processed_ids():
    base = os.path.join(ROOT, "POSTES UTILIZADOS")
    pat = re.compile(r"COD ID_(\d+)")
    ids = set()
    for sub in list(_CATEGORY_SUBDIRS) + ["DUPLICADOS"]:
        d = os.path.join(base, sub)
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            m = pat.match(fn)
            if m:
                ids.add(m.group(1))
    return ids


def find_existing_files(cod_id):
    """Arquivos ja existentes pra esse cod_id nas filas normais (nao conta
    o que ja esta em DUPLICADOS -- esse ja foi tratado antes)."""
    base = os.path.join(ROOT, "POSTES UTILIZADOS")
    out = []
    for sub in _CATEGORY_SUBDIRS:
        d = os.path.join(base, sub)
        if not os.path.isdir(d):
            continue
        out.extend(glob.glob(os.path.join(d, "COD ID_%s*.jpg" % cod_id)))
    return out


def load_processados():
    """Abre (ou cria) a planilha-mestre de tudo que ja foi processado, com o
    link usado + o link oficial da base lado a lado pra conferencia. Fica
    aberta em memoria o lote inteiro e so salva uma vez no final (esse
    arquivo cresce bastante -- reabrir/salvar a cada linha ficaria lento)."""
    header = ["cod_id", "endereco", "pasta_atual", "link_usado", "link_original_base"]
    if os.path.exists(PROCESSADOS_XLSX):
        wb = openpyxl.load_workbook(PROCESSADOS_XLSX)
        ws = wb["processados"] if "processados" in wb.sheetnames else wb.active
        col = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
        for name in header:
            if name not in col:
                col[name] = ws.max_column + 1
                ws.cell(row=1, column=col[name], value=name).font = Font(bold=True)
        existing = set()
        for r in range(2, ws.max_row + 1):
            cod = ws.cell(row=r, column=col["cod_id"]).value
            if cod:
                existing.add(str(cod))
        return wb, ws, existing, col
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "processados"
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    col = {name: i + 1 for i, name in enumerate(header)}
    return wb, ws, set(), col


def append_xlsx_rows(path, sheet_title, header, rows):
    """Abre a planilha se ja existir, ou cria uma nova com `header`. `rows` e
    uma lista de DICTS {nome_da_coluna: valor} -- nunca posicional, porque o
    usuario pode ter inserido/reordenado coluna (achado 23/ago: apendar por
    posicao desalinhou tudo assim que o usuario acrescentou 'Revisado?' na
    frente de duplicados_para_validar.xlsx). Coluna que o dict nao menciona
    fica em branco na linha nova -- normal, e coisa que so o usuario
    preenche (tipo 'Revisado?')."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_title] if sheet_title in wb.sheetnames else wb.active
        col = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(header)
        for c in ws[1]:
            c.font = Font(bold=True)
        col = {name: i + 1 for i, name in enumerate(header)}

    for name in header:
        if name not in col:
            col[name] = ws.max_column + 1
            ws.cell(row=1, column=col[name], value=name).font = Font(bold=True)

    for row_dict in rows:
        new_row = ws.max_row + 1
        for name, value in row_dict.items():
            if name not in col:
                col[name] = ws.max_column + 1
                ws.cell(row=1, column=col[name], value=name).font = Font(bold=True)
            ws.cell(row=new_row, column=col[name], value=value)
        linkify_row(ws, new_row)
    wb.save(path)


def linkify_row(ws, row_idx):
    """Transforma em hyperlink de verdade toda celula da linha que comeca
    com http -- por padrao openpyxl so escreve texto puro, e o Excel so
    autolinca o que uma pessoa digita/cola direto, nao o que um script
    escreve (foi o que o usuario notou 23/ago)."""
    for cell in ws[row_idx]:
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


def dest_dir_for(cod_id, esquina_ids):
    """Sempre devolve a fila de revisao da categoria (nunca a '- OK' --
    essa so recebe arquivo quando um humano confirma e move a mao)."""
    base = os.path.join(ROOT, "POSTES UTILIZADOS")
    sub = "ESQUINA - REVISAO" if cod_id in esquina_ids else "POSTE UTILIZADO - REVISAO"
    d = os.path.join(base, sub)
    os.makedirs(d, exist_ok=True)
    return d


def match_cod_id(cam_lat, cam_lon, aim_h, poles, sv_address=None):
    cands = sorted(((hav(cam_lat, cam_lon, la, lo), cod, la, lo, addr) for cod, la, lo, addr in poles),
                    key=lambda t: t[0])[:6]
    scored = [(d, ang_diff(bearing_deg(cam_lat, cam_lon, la, lo), aim_h), cod, addr)
              for d, cod, la, lo, addr in cands]
    nearest = scored[0]
    best_bearing = min(scored, key=lambda t: t[1])

    flagged, note = False, ""
    if nearest[2] == best_bearing[2]:
        chosen = nearest
    elif nearest[0] < 10 and best_bearing[1] < 30:
        chosen = nearest
        if nearest[1] > 60:
            flagged, note = True, (
                "confirmar - poste mais proximo (%.1fm) tem %.0f graus de diferenca de "
                "rumo, mas esta perto o suficiente pra ser so ruido de GPS" % (nearest[0], nearest[1]))
    elif nearest[1] > 60 and best_bearing[0] >= 10 and best_bearing[1] < 30:
        chosen = best_bearing
        note = ("cod_id escolhido pelo rumo, nao pela distancia (mais proximo a %.1fm "
                 "tinha %.0f graus de diferenca de rumo)" % (nearest[0], nearest[1]))
    else:
        chosen = nearest
        flagged, note = True, (
            "confirmar - distancia (%.1fm, %.0f graus) e rumo (%.1fm, %.0f graus) nao "
            "concordam com confianca" % (nearest[0], nearest[1], best_bearing[0], best_bearing[1]))

    # checagem de nome de rua (achado 23/ago, refinada no mesmo dia depois de
    # rodar contra 218 casos reais: comparar nome EXATO gera falso positivo
    # demais -- "Kubitscheck" vs "Kubitschek", "Drumond" vs "Drummond", "Ver.
    # Jose Rosinha" vs "Vereador Jose Rosinha", postes de esquina onde a rua
    # mais proxima varia -- nada disso e erro de verdade. Por isso: NUNCA
    # troca o cod_id sozinho (risco de trocar um acerto por um erro e maior
    # que o beneficio); so usa `difflib` (similaridade aproximada, nao nome
    # exato) pra avisar quando a rua bate tao mal que quase certamente e erro
    # mesmo. O endereco vem de meta['address'] (Street View), sem custo extra.
    sv_street = normalize_street_name(sv_address) if sv_address else ""
    if sv_street and sv_street not in NO_STREET:
        chosen_street = normalize_street_name(chosen[3])
        sim = street_word_similarity(sv_street, chosen_street)
        if sim < 0.4:
            flagged = True
            note = (("%s; " % note) if note else "") + (
                "ATENCAO nome de rua muito diferente: Street View diz o endereco e '%s', "
                "cod_id escolhido e de '%s' - conferir manualmente (nao troquei sozinho, "
                "pode ser esquina ou so grafia diferente)" % (sv_address, chosen[3]))

    return chosen[2], chosen[0], chosen[1], flagged, note, chosen[3]


def audit_street_match(link_url, cod_endereco, poles, known_meta=None):
    """Pra enriquecer duplicados_para_validar.xlsx (pedido do usuario 23/ago):
    confere se a rua que o Street View reporta pro link bate com o endereco
    do cod_id que ele acabou levando, e sugere um candidato melhor por
    distancia+rua quando nao bate -- so informativo, nunca aplicado sozinho
    (mesma cautela do match_cod_id). Returns (sv_address, 'SIM'/'NAO'/'(sem
    dado de rua)', sugestao_cod_id_com_distancia, endereco_da_sugestao)."""
    if not link_url:
        return "", "", "", ""
    if known_meta is not None:
        meta = known_meta
    else:
        m = RE_MAIN.search(link_url)
        if not m:
            return "", "", "", ""
        meta = get_pano_meta(float(m.group(1)), float(m.group(2)))
    if not meta:
        return "", "", "", ""
    sv_addr = meta.get("address")
    sv_norm = normalize_street_name(sv_addr) if sv_addr else ""
    if not sv_norm or sv_norm in NO_STREET:
        return sv_addr or "", "(sem dado de rua)", "", ""
    if street_word_similarity(sv_norm, normalize_street_name(cod_endereco)) >= 0.4:
        return sv_addr, "SIM", "", ""
    cam_lat, cam_lon = meta.get("lat"), meta.get("lon")
    if cam_lat is None or cam_lon is None:
        return sv_addr, "NAO", "", ""
    cands = sorted(((hav(cam_lat, cam_lon, la, lo), cod, addr) for cod, la, lo, addr in poles),
                    key=lambda t: t[0])[:8]
    best = None
    for d, cod, addr in cands:
        if street_word_similarity(sv_norm, normalize_street_name(addr)) >= 0.4:
            if best is None or d < best[0]:
                best = (d, cod, addr)
    if best:
        return sv_addr, "NAO", "%s (dist=%.1fm)" % (best[1], best[0]), best[2]
    return sv_addr, "NAO", "", ""


def find_pole_x(arr):
    gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
    edges = cv2.Canny(gray, 50, 150)
    edges[:, :LEFT_UI_EXCLUDE] = 0
    edges[:, RIGHT_UI_EXCLUDE:] = 0
    h, w = edges.shape
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=80, minLineLength=int(h * 0.25), maxLineGap=15)
    if lines is None:
        return None
    # opencv-python 5.x mudou o retorno do HoughLinesP de (N,1,4) pra (N,4)
    # (achado 24/ago, rodando contra opencv 5.0.0 -- lines[:, 0, :] quebrava
    # com "too many indices" assim que uma linha era encontrada); reshape
    # normaliza os dois formatos sem custo (no-op no formato antigo).
    lines = lines.reshape(-1, 4)
    cx = (LEFT_UI_EXCLUDE + RIGHT_UI_EXCLUDE) / 2
    best_score, best_mid = 0.0, None
    for (x1, y1, x2, y2) in lines:
        dx, dy = x2 - x1, y2 - y1
        length = math.hypot(dx, dy)
        if length < 1:
            continue
        angle = abs(math.degrees(math.atan2(dx, dy)))
        if angle > 6:
            continue
        mid_x = (x1 + x2) / 2
        centrality = max(0.0, 1 - abs(mid_x - cx) / (w * 0.4))
        lowness = max(0.0, 1 - (min(y1, y2) / (h * 0.75)))
        score = length * (0.3 + 0.7 * centrality) * (0.4 + 0.6 * lowness)
        if score > best_score:
            best_score, best_mid = score, mid_x
    return best_mid


def main():
    wb = openpyxl.load_workbook(PLANILHA)
    ws = wb["links"] if "links" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    for needed in ["link", "Poste", "cod_id", "endereco", "Processado?", "precisa_confirmar", "motivo"]:
        if needed not in col:
            col[needed] = ws.max_column + 1
            ws.cell(row=1, column=col[needed], value=needed).font = Font(bold=True)

    pending = []
    full_row_by_cod = {}  # cod_id -> (linha, link, endereco) para QUALQUER linha ja processada
    for r in range(2, ws.max_row + 1):
        url = cell_url(ws.cell(row=r, column=col["link"]))
        done = ws.cell(row=r, column=col["Processado?"]).value
        cod = ws.cell(row=r, column=col["cod_id"]).value
        if cod and cod not in full_row_by_cod:
            full_row_by_cod[cod] = (r, url, ws.cell(row=r, column=col["endereco"]).value)
        if url and str(done).strip().upper() != "X":
            pending.append(r)

    if len(sys.argv) > 1:
        pending = pending[:int(sys.argv[1])]

    print("linhas pendentes:", len(pending))
    if not pending:
        print("nada a fazer.")
        return

    os.makedirs(DUPLICADOS_DIR, exist_ok=True)
    poles = load_poles()
    esquina_ids = load_esquina()
    base_links = load_base_links()
    poste_map = load_poste_numero_map()
    used_ids = already_processed_ids()  # already on disk from before
    proc_wb, proc_ws, proc_ids, proc_col = load_processados()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome",
                                     args=["--disable-blink-features=AutomationControlled"])
        context = browser.new_context(viewport=VIEWPORT, locale="pt-BR", device_scale_factor=1)
        page = context.new_page()
        page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        for r in pending:
            candidatos = link_candidates(ws.cell(row=r, column=col["link"]))
            url = candidatos[0] if candidatos else None
            t0 = time.time()
            try:
                # tenta cada candidato (valor direto = caminho tradicional,
                # depois hyperlink) e fica com o primeiro que parsear certo
                L = None
                for cand in candidatos:
                    L = parse_link(cand)
                    if L is not None:
                        url = cand
                        break
                if L is None:
                    # fallback (achado 26/ago): o link pode ter coordenada
                    # (@lat,lon) mas sem angulo/panoid valido -- link de
                    # "viewpoint" generico (sem 3a/y/h/t) ou link quebrado
                    # (falta o h, ou panoid veio vazio). Em vez de desistir,
                    # so quando a coluna Poste confirma sem ambiguidade QUAL
                    # poste mirar, reconstroi o angulo pelo rumo geometrico
                    # ate esse poste (mesma logica de yaw_from_bearing que o
                    # run_batch.py ja usa quando nao ha link nenhum) e monta
                    # uma URL nova a partir do panoid fresco daquela
                    # coordenada. Sem Poste preenchido nao tenta -- nao ha
                    # como saber qual poste mirar sem angulo nem numero de
                    # campo.
                    coords_fb = extrai_coords_fallback(candidatos)
                    poste_num_fb = ws.cell(row=r, column=col["Poste"]).value
                    poste_num_fb = str(poste_num_fb).strip() if poste_num_fb not in (None, "") else None
                    motivo_fallback = None
                    if coords_fb and poste_num_fb and poste_num_fb in poste_map:
                        lat_fb, lon_fb = coords_fb
                        meta_fb = get_pano_meta(lat_fb, lon_fb)
                        alvo_fb = next(((la, lo) for c, la, lo, a in poles if c == poste_map[poste_num_fb]), None)
                        if meta_fb and meta_fb.get("lat") is not None and alvo_fb:
                            bearing_fb = bearing_deg(meta_fb["lat"], meta_fb["lon"], alvo_fb[0], alvo_fb[1])
                            yaw_fb = yaw_from_bearing(bearing_fb, meta_fb["heading"])
                            url = (f"https://www.google.com/maps/@{lat_fb},{lon_fb},3a,80y,"
                                   f"{yaw_fb:.2f}h,82t/data=!3m6!1e1!3m4!1s{meta_fb['panoid']}"
                                   f"!2e0!7i16384!8i8192")
                            L = {"lat": lat_fb, "lon": lon_fb, "y": 80.0, "h": yaw_fb, "t": 82.0,
                                 "pitch": 90 - 82.0, "panoid": meta_fb["panoid"]}
                            print(r, "RECONSTRUIDO via rumo geometrico (Poste %s)" % poste_num_fb)
                        else:
                            motivo_fallback = ("link sem angulo/panoid valido; tentei reconstruir pelo "
                                                "Poste %s mas o Google nao tem cobertura Street View "
                                                "nessa coordenada" % poste_num_fb)
                if L is None:
                    ws.cell(row=r, column=col["motivo"],
                            value=motivo_fallback or "URL nao reconhecida (nao tem 3a,Ny,Nh,Nt)")
                    ws.cell(row=r, column=col["Processado?"], value="X")
                    print(r, "PARSE FAIL" if not motivo_fallback else "PARSE FAIL (sem cobertura)")
                    continue

                meta = get_pano_meta(L["lat"], L["lon"])
                if meta is None or meta.get("lat") is None:
                    ws.cell(row=r, column=col["motivo"], value="sem metadados/cobertura Street View")
                    ws.cell(row=r, column=col["Processado?"], value="X")
                    print(r, "SEM METADADOS")
                    continue

                cam_lat, cam_lon = meta["lat"], meta["lon"]

                # numero do poste preenchido a mao (coluna 'Poste', pedido do
                # usuario 23/ago pra esquina/postes proximos) e fonte da
                # verdade -- substitui o match por distancia/rumo/rua inteiro,
                # nao e so mais um sinal (poste_numero e 1:1 com cod_id).
                poste_num = ws.cell(row=r, column=col["Poste"]).value
                poste_num = str(poste_num).strip() if poste_num not in (None, "") else None

                if poste_num:
                    if poste_num in poste_map:
                        cod_id = poste_map[poste_num]
                        addr = next((a for c, la, lo, a in poles if c == cod_id), "")
                        dist = hav(cam_lat, cam_lon,
                                   *next(((la, lo) for c, la, lo, a in poles if c == cod_id), (cam_lat, cam_lon)))
                        diff = 0.0
                        flagged, note = False, ""
                    else:
                        cod_id, dist, diff, flagged, note, addr = match_cod_id(
                            cam_lat, cam_lon, L["h"], poles, sv_address=meta.get("address"))
                        flagged = True
                        note = (("%s; " % note) if note else "") + (
                            "ATENCAO: numero de poste '%s' preenchido na planilha nao foi "
                            "encontrado na base -- usei o match por distancia/rumo/rua como "
                            "recurso, confirmar" % poste_num)
                else:
                    cod_id, dist, diff, flagged, note, addr = match_cod_id(
                        cam_lat, cam_lon, L["h"], poles, sv_address=meta.get("address"))

                is_collision = cod_id in used_ids

                page.goto(url, wait_until="load", timeout=45000)
                page.wait_for_timeout(6000)
                raw_path = "planilha_raw_%d.png" % r
                page.screenshot(path=raw_path)
                img = Image.open(raw_path).convert("RGB")
                arr = np.asarray(img)
                h, w = arr.shape[:2]

                pole_x = find_pole_x(arr)
                if pole_x is None:
                    pole_x = (LEFT_UI_EXCLUDE + RIGHT_UI_EXCLUDE) / 2

                usable_h = h - TOP_MARGIN - BOTTOM_MARGIN
                crop_w = int(usable_h * (OUT_W / OUT_H))
                x0 = int(max(0, min(w - crop_w, pole_x - crop_w / 2)))
                x1 = x0 + crop_w
                crop = arr[TOP_MARGIN:h - BOTTOM_MARGIN, x0:x1]
                crop_img = Image.fromarray(crop).resize((OUT_W, OUT_H), Image.LANCZOS)
                clean, wm_hits, _ = remove_watermark_v3(np.asarray(crop_img))

                base_link = base_links.get(cod_id, "")

                DUP_HEADER = ["cod_id", "linha_planilha", "link_usado", "endereco", "link_original_base",
                              "arquivo", "endereco_street_view", "rua_bate_com_cod_id_atual?",
                              "melhor_candidato_por_rua", "endereco_melhor_candidato", "qual_e_o_correto?"]

                if is_collision:
                    # nunca fica na fila normal: manda os DOIS lados (o que ja
                    # existia + este novo) pra DUPLICADOS, sem comentario no
                    # nome -- so numero da linha da planilha pra rastrear.
                    dup_rows = []
                    for existing_path in find_existing_files(cod_id):
                        orig = full_row_by_cod.get(cod_id)
                        label = str(orig[0]) if orig else "lote anterior"
                        new_path = os.path.join(DUPLICADOS_DIR, "COD ID_%s (linha %s).jpg" % (cod_id, label))
                        if os.path.exists(new_path):
                            continue  # ja movido numa rodada anterior
                        shutil.move(existing_path, new_path)
                        orig_link = orig[1] if orig else ""
                        sv_a, match_a, alt_a, alt_addr_a = audit_street_match(orig_link, addr, poles)
                        dup_rows.append(dict(zip(DUP_HEADER, [
                            cod_id, label, orig_link, orig[2] if orig else "", base_link,
                            os.path.basename(new_path), sv_a, match_a, alt_a, alt_addr_a, ""])))
                    new_path = os.path.join(DUPLICADOS_DIR, "COD ID_%s (linha %d).jpg" % (cod_id, r))
                    Image.fromarray(clean).save(new_path, quality=92)
                    sv_b, match_b, alt_b, alt_addr_b = audit_street_match(
                        url, addr, poles, known_meta=meta)
                    dup_rows.append(dict(zip(DUP_HEADER, [
                        cod_id, r, url, addr, base_link, os.path.basename(new_path),
                        sv_b, match_b, alt_b, alt_addr_b, ""])))
                    append_xlsx_rows(DUPLICADOS_XLSX, "duplicados", DUP_HEADER, dup_rows)
                    note = "COLISAO DE COD_ID - movido pra pasta DUPLICADOS, ver duplicados_para_validar.xlsx"
                    flagged = True
                    dest_label = "DUPLICADOS"
                else:
                    # regra do usuario 23/ago: qualquer duvida (flagged) NUNCA
                    # cai na fila normal da categoria (ESQUINA/POSTE UTILIZADO
                    # - REVISAO) -- essas ficam reservadas pra match confiante.
                    # Toda duvida vai pra BAIXA CONFIANCA - REVISAO (duplicata
                    # ja vai pra DUPLICADOS, tratado acima).
                    if flagged:
                        dest_dir = os.path.join(ROOT, "POSTES UTILIZADOS", "BAIXA CONFIANCA - REVISAO")
                        os.makedirs(dest_dir, exist_ok=True)
                    else:
                        dest_dir = dest_dir_for(cod_id, esquina_ids)
                    name = "COD ID_%s%s.jpg" % (cod_id, "!!" if flagged else "")
                    Image.fromarray(clean).save(os.path.join(dest_dir, name), quality=92)
                    dest_label = os.path.basename(dest_dir)
                    if flagged:
                        append_xlsx_rows(
                            CONFIRMAR_XLSX, "confirmar",
                            ["cod_id", "linha_planilha", "link", "endereco", "motivo", "pasta"],
                            [dict(cod_id=cod_id, linha_planilha=r, link=url, endereco=addr,
                                  motivo=note, pasta=dest_label)],
                        )

                used_ids.add(cod_id)
                ws.cell(row=r, column=col["cod_id"], value=cod_id)
                ws.cell(row=r, column=col["endereco"], value=addr)
                ws.cell(row=r, column=col["precisa_confirmar"], value="SIM" if flagged else "")
                ws.cell(row=r, column=col["motivo"], value=note)
                ws.cell(row=r, column=col["Processado?"], value="X")

                if cod_id not in proc_ids:
                    new_row = proc_ws.max_row + 1
                    for name, value in dict(cod_id=cod_id, endereco=addr, pasta_atual=dest_label,
                                             link_usado=url, link_original_base=base_link).items():
                        proc_ws.cell(row=new_row, column=proc_col[name], value=value)
                    linkify_row(proc_ws, new_row)
                    proc_ids.add(cod_id)

                print(r, cod_id, "!!" if flagged else "", "->", dest_label,
                      "dist=%.1fm diff=%.0f (%.1fs)" % (dist, diff, time.time() - t0))
            except Exception as e:
                ws.cell(row=r, column=col["motivo"], value="ERRO: %s" % e)
                print(r, "ERRO", e)
            finally:
                # save incrementally so progress isn't lost if a later row fails
                # hard -- precisa ser 'finally', nao so uma linha solta apos o
                # try/except: os 'continue' de PARSE FAIL / SEM METADADOS estao
                # dentro do try e pulavam direto pra proxima iteracao, nunca
                # chegando a salvar essa linha em disco. Se as ultimas linhas da
                # planilha caissem todas nesses ramos (sem nenhuma linha normal
                # depois pra "carregar" o save pendente), o X+motivo delas
                # ficava so na memoria e se perdia ao fechar o script (achado
                # 25/ago, 8 linhas no fim da planilha nunca gravavam).
                wb.save(PLANILHA)

        browser.close()

    proc_wb.save(PROCESSADOS_XLSX)
    print("atualizado:", PROCESSADOS_XLSX)
    print("\nDONE.")


if __name__ == "__main__":
    main()
